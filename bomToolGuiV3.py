from tkinter import ttk,filedialog
from tkinter import *
import pandas as pd
# import argparse
from openpyxl import Workbook,worksheet
from openpyxl.styles import Border, Side, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

root = Tk()
root.title('eagel bom tool')

isOctoPart = BooleanVar()
isOctoPart.set(False)

octoPartUser = StringVar()
octoPartUser.set('jhonedoe@gmail.com')

pathLabelText = 'Select file'
path = ''

def selectFile():
    global path
    path  = filedialog.askopenfilename(initialdir = "`~`",title = "Select file",filetypes = (("csv files","*.csv"),("all files","*.*")))
    if len(path) > 0:
        print(path)
        pathEntry.delete(0, END)
        pathEntry.insert(END, path)

def get_col_widths(dataframe):
    # First we find the maximum length of the index column   
    idx_max = max([len(str(s)) for s in dataframe.index.values] + [len(str(dataframe.index.name))])
    # Then, we concatenate this to the max of the lengths of column name and its values for each column, left to right
    return [idx_max] + [max([len(str(s)) for s in dataframe[col].values] + [len(col)]) for col in dataframe.columns]

def runScript():
    filePath = pathEntry.get()
    print(filePath)
    if isOctoPart.get() == True:
        print("bulding an Octopart BOM")
    if len(filePath) > 0:
        cuurencyFormat = '#,##0.00$'
        #Import CSV file as dataframe
        bomDf = pd.read_csv(filePath,sep=';')
        #Remove any exluded parts from dataframe
        if 'BOM' in bomDf.columns:
            indexNames = bomDf[bomDf['BOM'] == 'EXCLUDE'].index
            bomDf.drop(indexNames, inplace=True)
        #copy the importent columes to a new dataframe
        newDf = pd.DataFrame()
        newDf = bomDf[['Qty', 'Parts','MF','MPN','VALUE','FOOTPRINT','DESCRIPTION']].copy()
        #creat a new workbook 
        wb = Workbook()
        ws = wb.active
        #copy new dataframe to workbook
        for r in dataframe_to_rows(newDf, index=False, header=True):
            ws.append(r)
        #delete blank row
        ws.delete_rows(2,1)
        #add index column 
        ws.insert_cols(1,1)
        ws.cell(1,1).value = '#'
        for row in range(2, ws.max_row+1):
            ws.cell(row,1).value = row-1
        #save space for octopart columns
        if isOctoPart.get() == True:   
            ws.insert_cols(6,2)
            ws.cell(1,6,'PRICE')
            ws.cell(1,7,'TOTAL')
        #set column width
        for i, width in enumerate(get_col_widths(newDf)):
            if (i != 2):
                if isOctoPart.get() == True:
                    if(i < 6):
                        ws.column_dimensions[get_column_letter(i+1)].width = width
                    else:
                        ws.column_dimensions[get_column_letter(i+3)].width = width
                else:
                    ws.column_dimensions[get_column_letter(i+1)].width = width
            else:
                 ws.column_dimensions[get_column_letter(i+1)].width = 20
                         
        #add space for BOM title
        ws.insert_rows(0,3)
        #marge hader row
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
        #build the hader string 
        name_index = filePath.rfind("/")
        title_string = filePath[(name_index+1):]
        title_string = title_string.replace('.csv',' BOM')
        #Writing the hader and adding style
        ws.cell(row = 1, column = 1).value = title_string
        ws.cell(row = 1, column = 1).font = Font(size = 22, bold = True, underline='single') 
        ws.cell(row = 1, column = 1).alignment = Alignment(horizontal='center',vertical='center')
      
        #if Octopart BOM
        if isOctoPart.get() == True:
            priceFormula = '=OCTOPART_DISTRIBUTOR_PRICE(%s,%s,"digikey",1000)'
            for row in range(5, ws.max_row+1):
                qtyCell = ws.cell(row,2).coordinate
                mpnCell = ws.cell(row,4).coordinate
                mfCell = ws.cell(row,5).coordinate
                formula = priceFormula % (mfCell,mpnCell)
                ws.cell(row,6).value = formula
                ws.cell(row,6).number_format = cuurencyFormat
                currCell = ws.cell(row,6).coordinate
                formula = '=%s*%s' % (qtyCell,currCell)
                ws.cell(row,7).value = formula
                ws.cell(row,7).number_format = cuurencyFormat

        #set borders
        cell_border = Border(left=Side(border_style='thin', color='FF000000'),
                            right=Side(border_style='thin', color='FF000000'),
                            top=Side(border_style='thin', color='FF000000'),
                            bottom=Side(border_style='thin', color='FF000000'))
     
        for row in range(4, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row, col)
                cell.border = cell_border
                #cell.alignment = alignment
                if row == 4:        # Header Style
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center',vertical='center')
                elif col < 3:
                    cell.alignment = Alignment(horizontal='center',vertical='center')
                elif col == 3:
                    cell.alignment = Alignment(horizontal='left',vertical='top',wrapText=True)
                else:
                    cell.alignment = Alignment(horizontal='left',vertical='top')

        ws.page_setup.fitToWidth = 1
        
        #Rename the file path for the ne file
        if isOctoPart.get() == False:
            newFileName = filePath.replace('.csv','_BOM.xlsx')
        else:
            start = ws.cell(5,7).coordinate
            end = ws.cell(ws.max_row,7).coordinate
            ws.cell(row = ws.max_row+1, column = 7).value = '=SUM(%s:%s)' % (start,end)
            ws.cell(row = ws.max_row, column = 7).number_format = cuurencyFormat
            ws.cell(row = ws.max_row, column = 7).alignment= Alignment(horizontal='center',vertical='center')
            ws.cell(row = ws.max_row, column = 1).value = '=OCTOPART_SET_USER("%s")' % octoPartUser.get()
            newFileName = filePath.replace('.csv','_BOM_octopart.xlsx')
        #save the new excel file
        wb.save(newFileName)
       
        pathEntry.delete(0, END)
        pathEntry.insert(END, 'done select new file')

        

pathLabel = ttk.Label(root, text=pathLabelText, background="white").grid(row= 4, column=0,columnspan=3, padx=10, pady=2, sticky="W")

pathEntry = ttk.Entry(root, width=60, background="gray")
pathEntry.grid(row= 5, column=0, columnspan=5, padx=10, pady=4, sticky="WE")

fileSelectBtn = Button(root, text="Select a file", command=selectFile).grid(row= 5, column=6, padx=5, pady=4, sticky="W")
runScriptBtn = Button(root, text="RUN", command=runScript).grid(row= 5, column=7, padx=5, pady=4,sticky="W")
octoPartBtn = Checkbutton(root, text="Octopart", variable=isOctoPart, onvalue = 1, offvalue = 0).grid(row= 6, column=0, padx=5, pady=4,sticky="W")
#fileSelectBtn.pack(side="bottom", fill="both", expand="yes", padx="10", pady="10")

root.mainloop()
